import os
import json
import xml.etree.ElementTree as ET
import requests
import logging
import hashlib
from dotenv import load_dotenv
from datetime import datetime, timedelta
from apscheduler.schedulers.blocking import BlockingScheduler
from contextlib import contextmanager
import threading
import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime

# Carrega variáveis de ambiente do .env
load_dotenv()

class Config:
    def __init__(self):
        self.zig_token = os.getenv('ZIG_TOKEN')
        self.zig_rede = os.getenv('ZIG_REDE')
        self.zig_rede2 = os.getenv('ZIG_REDE2')
        self.omie_app_key = os.getenv('OMIE_APP_KEY')
        self.omie_app_secret = os.getenv('OMIE_APP_SECRET')

class NFeProc:
    def __init__(self, xml_data):
        self.root = ET.fromstring(xml_data)
        self.inf_nfe = self.root.find(".//infNFe")
        self.id = self.inf_nfe.attrib['Id']
        self.ide = self.inf_nfe.find("ide")
        self.det = self.inf_nfe.findall("det")
        self.total = self.inf_nfe.find("total/ICMSTot")
        self.emit = self.inf_nfe.find("emit")
        self.nfce = None

class SimplifiedInvoice:
    def __init__(self, nf_number, product_name, value, impostos):
        self.nf_number = nf_number
        self.product_name = product_name
        self.value = value
        self.impostos = impostos

config = Config()
scheduler = BlockingScheduler()

# Timeout usando threading.Timer
@contextmanager
def timeout(duration):
    def raise_timeout():
        raise TimeoutError("Operation timed out.")
    
    timer = threading.Timer(duration, raise_timeout)
    timer.start()
    try:
        yield
    finally:
        timer.cancel()

def fetch_invoices(from_date, to_date, page):
    headers = {
        "Authorization": config.zig_token,
    }
    params = {
        #"dtinicio": from_date.strftime('%Y-%m-%d'),
        #"dtfim": to_date.strftime('%Y-%m-%d'),
        "dtinicio": from_date.strftime('2024-10-13'),
        "dtfim": to_date.strftime('2024-10-15'),
        "loja": config.zig_rede2,
        "page": str(page)
    }
    response = requests.get("https://api.zigcore.com.br/integration/erp/invoice?", headers=headers, params=params)
    
    if response.status_code != 200:
        raise Exception(f"Unexpected status: {response.status_code}")
    
    return response.json()
def create_xlsx_from_omie_json(omie_json, filename=None):
    if filename is None:
        filename = f"omie_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dados Omie"

    # Cabeçalho
    headers = ["Chave NF-e", "Data Emissão", "Hora Emissão", "Número NF", "Série", "Ambiente", "Tipo Emissão", "Valor Total"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Dados gerais da NF-e
    nfe_data = omie_json["NFe"]
    ws.append([
        nfe_data["chNFe"],
        nfe_data["dEmi"],
        nfe_data["hEmi"],
        nfe_data["nNF"],
        nfe_data["serie"],
        "Produção" if nfe_data["tpAmb"] == "P" else "Homologação",
        nfe_data["tpEmis"],
        nfe_data["total"]["vCF"]
    ])

    # Adicionar uma linha em branco
    ws.append([])

    # Cabeçalho dos itens
    item_headers = ["Sequência", "Código", "Descrição", "NCM", "CFOP", "Unidade", "Quantidade", "Valor Unitário", "Valor Total"]
    for col, header in enumerate(item_headers, 1):
        cell = ws.cell(row=ws.max_row + 1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Dados dos itens
    for item in nfe_data["det"]:
        ws.append([
            item["seqItem"],
            item["prod"]["cProd"],
            item["prod"]["xProd"],
            item["prod"]["NCM"],
            item["prod"]["CFOP"],
            item["prod"]["cUn"],
            item["prod"]["nQuant"],
            item["prod"]["vUnit"],
            item["prod"]["vProd"]
        ])

    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(filename)
    return filename
def convert_xml_to_json(xml_string):
    def _xml_to_dict(element):
        result = {}
        for child in element:
            child_data = _xml_to_dict(child)
            if child.tag in result:
                if isinstance(result[child.tag], list):
                    result[child.tag].append(child_data)
                else:
                    result[child.tag] = [result[child.tag], child_data]
            else:
                result[child.tag] = child_data
        if element.attrib:
            result['@attributes'] = element.attrib
        if element.text and element.text.strip():
            if not result:
                return element.text.strip()
            else:
                result['#text'] = element.text.strip()
        return result

    root = ET.fromstring(xml_string)
    return json.dumps(_xml_to_dict(root), indent=2)

def convert_xml_to_omie_json(xml_data):
    nfe_json, nfe_data = parse_nfe_xml(xml_data)   
    omie_json = {
        "NFe": {
            "chNFe": nfe_data['Id'][3:] if nfe_data.get('Id') else "",
            "dEmi": datetime.strptime(nfe_data['ide']['dhEmi'], "%Y-%m-%dT%H:%M:%S%z").strftime("%d/%m/%Y"),
            "hEmi": datetime.strptime(nfe_data['ide']['dhEmi'], "%Y-%m-%dT%H:%M:%S%z").strftime("%H:%M:%S"),
            "nNF": nfe_data['ide']['nNF'],
            "serie": nfe_data['ide']['serie'],
            "tpAmb": "P" if nfe_data['ide']['tpAmb'] == "1" else "H",
            "tpEmis": nfe_data['ide']['tpEmis'],
            "lCanc": False,  # Assuming not cancelled
            "det": [],
            "total": {
                "vAcresc": "0.00",
                "vCF": nfe_data['total']['vNF'],
                "vDesc": nfe_data['total']['vDesc'],
                "vICMS": nfe_data['total']['vICMS'],
                "vItem": nfe_data['total']['vProd'],
                "vTaxa": 0,
                "vTotTrib": nfe_data['total']['vTotTrib']
            }
        },
        "caixa": {
            "lCxAberto": True,
            "seqCaixa": 0,  # You'll need to provide this information
            "seqCupom": 0  # You'll need to provide this information
        },
        "cupomIdent": {
            "idCliente": 0,  # You'll need to provide this information
            "idProjeto": 0,
            "idVendedor": 0
        },
        "emissor": {
            "emiId": 6029653,
            "emiNome": nfe_data['emit']['xFant'],
            "emiSerial": "",  # You'll need to provide this information
            "emiVersao": nfe_data['ide']['verProc']
        },
        "formasPag": [
             {
            "Parcelas": [
            ],
            "TEF": {
            },
            "lCanc": False,
            "lNaoGerarTitulo":False,
            "pag": {
                "pTaxa": 0,
                "vLiq": nfe_data['total']['vTotTrib'],
                "vPag": nfe_data['total']['vTotTrib'],
                "vTaxa": 0,
                "vTroco": 0
            },
            "pagIdent": {
                "cCategoria": "1.01.03",
                "cTipoPag": "99999",
                "idConta": 6758280882
            },
            "seqPag": 1
            }],  # You'll need to provide this information
        "nfce": {
            "nfceMd5": hashlib.md5(xml_data.encode()).hexdigest(),
            "nfceProt": nfe_data['nProt'],  # You'll need to provide this information
            "nfceXml": xml_data,
        }
    }
    
    for item in nfe_data['det']:
        det_item = {
            "lCanc": False,
            "lNaoMovEstoque": False,
            "prod": {
                "CFOP": item['prod']['CFOP'],
                "NCM": item['prod']['NCM'],
                "cEAN": item['prod']['cEAN'],
                "cProd": item['prod']['cProd'],
                "cUn": item['prod']['uCom'],
                "nQuant": float(item['prod']['qCom']),
                "vAcresc": 0,
                "vDesc": 0,
                "vItem": float(item['prod']['vProd']),
                "vProd": float(item['prod']['vProd']),
                "vUnit": float(item['prod']['vUnCom']),
                "xProd": item['prod']['xProd']
            },
            "prodIdent": {
                "emiProduto": item['prod']['cProd'],
                "idLocalEstoque": "",  # You'll need to provide this information
                "idProduto": 13  # You'll need to provide this information
            },
            "seqItem": int(item['nItem'])
        }
        omie_json["NFe"]["det"].append(det_item)    
    return omie_json

def build_omie_json(invoice):
    omie_json = convert_xml_to_omie_json(invoice["xml"])
    root = ET.fromstring(invoice["xml"])
    nfce_element = root.find('nfce')
    if nfce_element is not None:
        nfce_content = nfce_element.text.strip()
    # Add payment information
    #omie_json["formasPag"] = invoice["formasPag"]
    
    omie_json["caixa"]["seqCaixa"] = get_next_sequencial('seqCaixa')
    omie_json["caixa"]["seqCupom"] = get_next_sequencial('seqCupom')
    omie_json["cupomIdent"]["idCliente"] = invoice.get("idCliente", '675944858')
    omie_json["emissor"]["emiSerial"] = invoice.get("emiSerial", 1)
    #omie_json["nfce"]["nfceProt"] = nfe_data['nProt']
    
    return omie_json
def create_json_from_omie_json(omie_json, filename=None):
    if filename is None:
        filename = f"omie_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    
    # Cria uma cópia do omie_json para não modificar o original
    json_data = omie_json.copy()
    
    # Adiciona um timestamp ao JSON
    json_data['timestamp'] = datetime.now().isoformat()
    
    # Escreve o JSON em um arquivo
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(json_data, f, ensure_ascii=False, indent=4)
    
    return filename
def process_omie_invoice(omie_json):
    url = "https://app.omie.com.br/api/v1/produtos/cupomfiscalincluir/"
    headers = {"Content-Type": "application/json"}
    body = {
        "call": "IncluirNfce",
        "app_key": config.omie_app_key,
        "app_secret": config.omie_app_secret,
        "param": [omie_json]
    }
    response = requests.post(url, headers=headers, json=body)
    logging.info(f"JSON sendo enviado ao Omie: {json.dumps(omie_json, indent=2)}")
    if response.status_code != 200:
        raise Exception(f"Unexpected status: {response.status_code}")
    
    logging.info(f"Nota fiscal enviada com sucesso: {response.text}")

def execute_zig_omie_integration():
    logging.info("Iniciando integração...")
    
    now = datetime.now()
    last_run = now - timedelta(days=1)

    try:
        with timeout(240):  # Timeout de 4 minutos
            invoices = fetch_invoices(last_run, now, 1)
            for invoice in invoices:
                omie_json = build_omie_json(invoice)
                process_omie_invoice(omie_json)
    except TimeoutError:
        logging.error("O tempo limite foi atingido.")
    except Exception as e:
        logging.error(f"Erro na integração: {e}")

    logging.info("Processamento concluído.")
def parse_nfe_xml(xml_data):
    # Define o namespace
    ns = {'nfe': 'http://www.portalfiscal.inf.br/nfe', 'ds': 'http://www.w3.org/2000/09/xmldsig#'}

    # Carrega o XML
    root = ET.fromstring(xml_data)

    # Inicializa um dicionário para armazenar os dados
    nfe_data = {}

    # Encontra o elemento 'infNFe' com o namespace
    inf_nfe = root.find('.//nfe:infNFe', ns)
    
    if inf_nfe is not None:
        # Atributos de 'infNFe'
        nfe_data['Id'] = inf_nfe.attrib.get('Id')
        nfe_data['versao'] = inf_nfe.attrib.get('versao')

        # 'ide' - Identificação da NF-e
        ide = inf_nfe.find('nfe:ide', ns)
        if ide is not None:
            nfe_data['ide'] = {
                'cUF': ide.find('nfe:cUF', ns).text,
                'cNF': ide.find('nfe:cNF', ns).text,
                'natOp': ide.find('nfe:natOp', ns).text,
                'mod': ide.find('nfe:mod', ns).text,
                'serie': ide.find('nfe:serie', ns).text,
                'nNF': ide.find('nfe:nNF', ns).text,
                'dhEmi': ide.find('nfe:dhEmi', ns).text,
                'tpNF': ide.find('nfe:tpNF', ns).text,
                'idDest': ide.find('nfe:idDest', ns).text,
                'cMunFG': ide.find('nfe:cMunFG', ns).text,
                'tpImp': ide.find('nfe:tpImp', ns).text,
                'tpEmis': ide.find('nfe:tpEmis', ns).text,
                'cDV': ide.find('nfe:cDV', ns).text,
                'tpAmb': ide.find('nfe:tpAmb', ns).text,
                'finNFe': ide.find('nfe:finNFe', ns).text,
                'indFinal': ide.find('nfe:indFinal', ns).text,
                'indPres': ide.find('nfe:indPres', ns).text,
                'indIntermed': ide.find('nfe:indIntermed', ns).text,
                'procEmi': ide.find('nfe:procEmi', ns).text,
                'verProc': ide.find('nfe:verProc', ns).text
            }

        # 'emit' - Emitente
        emit = inf_nfe.find('nfe:emit', ns)
        if emit is not None:
            enderEmit = emit.find('nfe:enderEmit', ns)
            nfe_data['emit'] = {
                'CNPJ': emit.find('nfe:CNPJ', ns).text,
                'xNome': emit.find('nfe:xNome', ns).text,
                'xFant': emit.find('nfe:xFant', ns).text,
                'IE': emit.find('nfe:IE', ns).text,
                'CRT': emit.find('nfe:CRT', ns).text,
                'enderEmit': {
                    'xLgr': enderEmit.find('nfe:xLgr', ns).text,
                    'nro': enderEmit.find('nfe:nro', ns).text,
                    'xBairro': enderEmit.find('nfe:xBairro', ns).text,
                    'cMun': enderEmit.find('nfe:cMun', ns).text,
                    'xMun': enderEmit.find('nfe:xMun', ns).text,
                    'UF': enderEmit.find('nfe:UF', ns).text,
                    'CEP': enderEmit.find('nfe:CEP', ns).text,
                    'cPais': enderEmit.find('nfe:cPais', ns).text,
                    'xPais': enderEmit.find('nfe:xPais', ns).text,
                    'fone': enderEmit.find('nfe:fone', ns).text
                }
            }

        # 'dest' - Destinatário
        dest = inf_nfe.find('nfe:dest', ns)
        if dest is not None:
            nfe_data['dest'] = {
                'CPF': dest.find('nfe:CPF', ns).text,
                'xNome': dest.find('nfe:xNome', ns).text,
                'indIEDest': dest.find('nfe:indIEDest', ns).text
            }

        # 'det' - Detalhes dos produtos/serviços
        det_list = []
        for det in inf_nfe.findall('nfe:det', ns):
            prod = det.find('nfe:prod', ns)
            imposto = det.find('nfe:imposto', ns)
            det_item = {
                'nItem': det.attrib.get('nItem'),
                'prod': {
                    'cProd': prod.find('nfe:cProd', ns).text,
                    'cEAN': prod.find('nfe:cEAN', ns).text,
                    'xProd': prod.find('nfe:xProd', ns).text,
                    'NCM': prod.find('nfe:NCM', ns).text,
                    'CFOP': prod.find('nfe:CFOP', ns).text,
                    'uCom': prod.find('nfe:uCom', ns).text,
                    'qCom': prod.find('nfe:qCom', ns).text,
                    'vUnCom': prod.find('nfe:vUnCom', ns).text,
                    'vProd': prod.find('nfe:vProd', ns).text,
                    'cEANTrib': prod.find('nfe:cEANTrib', ns).text,
                    'uTrib': prod.find('nfe:uTrib', ns).text,
                    'qTrib': prod.find('nfe:qTrib', ns).text,
                    'vUnTrib': prod.find('nfe:vUnTrib', ns).text,
                    'indTot': prod.find('nfe:indTot', ns).text
                },
                'imposto': {
                    'vTotTrib': imposto.find('nfe:vTotTrib', ns).text if imposto.find('nfe:vTotTrib', ns) is not None else None,
                    # Aqui você pode adicionar mais detalhes do imposto se necessário
                }
            }
            det_list.append(det_item)
        nfe_data['det'] = det_list

        # 'total' - Totais da NF-e
        total = inf_nfe.find('nfe:total', ns)
        if total is not None:
            icms_tot = total.find('nfe:ICMSTot', ns)
            nfe_data['total'] = {
                'vBC': icms_tot.find('nfe:vBC', ns).text,
                'vICMS': icms_tot.find('nfe:vICMS', ns).text,
                'vICMSDeson': icms_tot.find('nfe:vICMSDeson', ns).text,
                'vFCP': icms_tot.find('nfe:vFCP', ns).text,
                'vBCST': icms_tot.find('nfe:vBCST', ns).text,
                'vST': icms_tot.find('nfe:vST', ns).text,
                'vFCPST': icms_tot.find('nfe:vFCPST', ns).text,
                'vFCPSTRet': icms_tot.find('nfe:vFCPSTRet', ns).text,
                'vProd': icms_tot.find('nfe:vProd', ns).text,
                'vFrete': icms_tot.find('nfe:vFrete', ns).text,
                'vSeg': icms_tot.find('nfe:vSeg', ns).text,
                'vDesc': icms_tot.find('nfe:vDesc', ns).text,
                'vII': icms_tot.find('nfe:vII', ns).text,
                'vIPI': icms_tot.find('nfe:vIPI', ns).text,
                'vIPIDevol': icms_tot.find('nfe:vIPIDevol', ns).text,
                'vPIS': icms_tot.find('nfe:vPIS', ns).text,
                'vCOFINS': icms_tot.find('nfe:vCOFINS', ns).text,
                'vOutro': icms_tot.find('nfe:vOutro', ns).text,
                'vNF': icms_tot.find('nfe:vNF', ns).text,
                'vTotTrib': icms_tot.find('nfe:vTotTrib', ns).text
            }
        prot_nfe = root.find('.//nfe:infProt', ns)
        if prot_nfe is not None:
            nfe_data['nProt'] = prot_nfe.find('nfe:nProt', ns).text if prot_nfe.find('nfe:nProt', ns) is not None else None

        # Outros elementos podem ser adicionados conforme necessário

    else:
        print("Elemento 'infNFe' não encontrado no XML.")

    # Converte o dicionário para JSON
    nfe_json = json.dumps(nfe_data, indent=4, ensure_ascii=False)
    return nfe_json, nfe_data  # Retorna o JSON e o dicionário de dados
def get_next_sequencial(tipo):
    filename = 'sequenciais.json'
    today = datetime.now().strftime('%Y-%m-%d')
    
    try:
        with open(filename, 'r') as f:
            data = json.load(f)
    except FileNotFoundError:
        data = {}

    if today not in data:
        data[today] = {'seqCaixa': 0, 'seqCupom': 0}

    data[today][tipo] += 1
    next_seq = data[today][tipo]

    with open(filename, 'w') as f:
        json.dump(data, f)

    return next_seq

if __name__ == "__main__":
    scheduler.add_job(execute_zig_omie_integration, 'interval', seconds=10)
    scheduler.start()
