import os
import json
import xml.etree.ElementTree as ET
import requests
import logging
import hashlib
import html
from dotenv import load_dotenv
from datetime import datetime, timedelta
from apscheduler.schedulers.blocking import BlockingScheduler
from contextlib import contextmanager
import threading
import openpyxl
import time
from openpyxl.styles import Font, Alignment
from datetime import datetime

# Carrega variáveis de ambiente do .env
load_dotenv()

class StoreConfig:
    def __init__(self, name, zig_token, zig_rede, omie_app_key, omie_app_secret,cc):
        self.name = name
        self.zig_token = zig_token
        self.zig_rede = zig_rede
        self.omie_app_key = omie_app_key
        self.omie_app_secret = omie_app_secret
        self.cc= cc
class Config:
    def __init__(self):
        self.stores = {
            'otro': StoreConfig(
                'otro',
                os.getenv('ZIG_TOKEN-OTRO'),
                os.getenv('ZIG_REDE-OTRO'),
                os.getenv('OMIE_APP_KEY-OTRO'),
                os.getenv('OMIE_APP_SECRET-OTRO'),
                os.getenv('CC-OTRO')
            ) ,
            'tratto': StoreConfig(
                'tratto',
                os.getenv('ZIG_TOKEN-TRATTO'),
                os.getenv('ZIG_REDE-TRATTO'),
                os.getenv('OMIE_APP_KEY-TRATTO'),
                os.getenv('OMIE_APP_SECRET-TRATTO'),
                os.getenv('CC-TRATTO')
            )
        }

config = Config()
scheduler = BlockingScheduler()

# Timeout usando threading.Timer
@contextmanager
def timeout(duration):
    def raise_timeout():
        raise TimeoutError("Operation timed out.")
    
    timer = threading.Timer(duration, raise_timeout)
    timer.start()
    try:
        yield
    finally:
        timer.cancel()

def fetch_invoices(store_config, from_date, to_date, page):
    headers = {
        "Authorization": store_config.zig_token,
    }
    params = {
        "dtinicio": from_date.strftime('%Y-%m-%d'),
        "dtfim": to_date.strftime('%Y-%m-%d'),
        #"dtinicio": from_date.strftime('2024-12-02'),
        #"dtfim": to_date.strftime('2024-10-23'),
        "loja": store_config.zig_rede,
        "page": str(page)
    }
    response = requests.get("https://api.zigcore.com.br/integration/erp/invoice?", headers=headers, params=params)
    if response.status_code != 200:
        raise Exception(f"Unexpected status: {response.status_code}")
    
    return response.json()
def create_xlsx_from_omie_json(omie_json, filename=None):
    if filename is None:
        filename = f"omie_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dados Omie"

    # Cabeçalho
    headers = ["Chave NF-e", "Data Emissão", "Hora Emissão", "Número NF", "Série", "Ambiente", "Tipo Emissão", "Valor Total"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Dados gerais da NF-e
    nfe_data = omie_json["NFe"]
    ws.append([
        nfe_data["chNFe"],
        nfe_data["dEmi"],
        nfe_data["hEmi"],
        nfe_data["nNF"],
        nfe_data["serie"],
        "Produção" if nfe_data["tpAmb"] == "P" else "Homologação",
        nfe_data["tpEmis"],
        nfe_data["total"]["vCF"]
    ])

    # Adicionar uma linha em branco
    ws.append([])

    # Cabeçalho dos itens
    item_headers = ["Sequência", "Código", "Descrição", "NCM", "CFOP", "Unidade", "Quantidade", "Valor Unitário", "Valor Total"]
    for col, header in enumerate(item_headers, 1):
        cell = ws.cell(row=ws.max_row + 1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Dados dos itens
    for item in nfe_data["det"]:
        ws.append([
            item["seqItem"],
            item["prod"]["cProd"],
            item["prod"]["xProd"],
            item["prod"]["NCM"],
            item["prod"]["CFOP"],
            item["prod"]["cUn"],
            item["prod"]["nQuant"],
            item["prod"]["vUnit"],
            item["prod"]["vProd"]
        ])

    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(filename)
    return filename
def convert_xml_to_json(xml_string):
    def _xml_to_dict(element):
        result = {}
        for child in element:
            child_data = _xml_to_dict(child)
            if child.tag in result:
                if isinstance(result[child.tag], list):
                    result[child.tag].append(child_data)
                else:
                    result[child.tag] = [result[child.tag], child_data]
            else:
                result[child.tag] = child_data
        if element.attrib:
            result['@attributes'] = element.attrib
        if element.text and element.text.strip():
            if not result:
                return element.text.strip()
            else:
                result['#text'] = element.text.strip()
        return result

    root = ET.fromstring(xml_string)
    return json.dumps(_xml_to_dict(root), indent=2)

def convert_xml_to_omie_json(xml_data):
    nfe_json, nfe_data = parse_nfe_xml(xml_data)
    nome_transformado = nfe_data['emit']['xFant'].replace("COMERCIO DE ", "").replace(" LTDA", "")
    xml_unescaped = html.unescape(xml_data)
    omie_json = {
        "NFe": {
            "chNFe": nfe_data['Id'][3:] if nfe_data.get('Id') else "",
            "dEmi": datetime.strptime(nfe_data['ide']['dhEmi'], "%Y-%m-%dT%H:%M:%S%z").strftime("%d/%m/%Y"),
            "hEmi": datetime.strptime(nfe_data['ide']['dhEmi'], "%Y-%m-%dT%H:%M:%S%z").strftime("%H:%M:%S"),
            "nNF": nfe_data['ide']['nNF'],
            "serie": nfe_data['ide']['serie'],
            "tpAmb": "P" if nfe_data['ide']['tpAmb'] == "1" else "H",
            "tpEmis": nfe_data['ide']['tpEmis'],
            "lCanc": False,  # Assuming not cancelled
            "det": [],
            "total": {
                "vAcresc": "0.00",
                "vCF": nfe_data['total']['vNF'],
                "vDesc": nfe_data['total']['vDesc'],
                "vICMS": nfe_data['total']['vICMS'],
                "vItem": nfe_data['total']['vProd'],
                "vTaxa": 0,
                "vTotTrib": nfe_data['total']['vTotTrib']
            }
        },
        "caixa": {
            "lCxAberto": False,
            "seqCaixa": 0,  # You'll need to provide this information
            "seqCupom": 0  # You'll need to provide this information
        },
        "cupomIdent": {
            "idCliente": 0,  # You'll need to provide this information
            "idProjeto": 0,
            "idVendedor": 0
        },
        "emissor": {
            "emiId": 6029653,
            "emiNome": nome_transformado,
            "emiSerial": "",  # You'll need to provide this information
            "emiVersao": nfe_data['ide']['verProc']
        },
        "formasPag": [
             {
            "Parcelas": [
            ],
            "TEF": {
            },
            "lCanc": False,
            "lNaoGerarTitulo":False,
            "pag": {
                "pTaxa": 0,
                "vLiq": nfe_data['total']['vNF'],  # Corrigido para o valor total correto
                "vPag": nfe_data['total']['vNF'],  # Certifique-se que o valor está correto
                "vTaxa": 0,
                "vTroco": 0
            },
            "pagIdent": {
                "cCategoria": " 1.01.03",
                "cTipoPag": "DIN",
                #"idConta":  7502625278
                "idConta": 0
            },
            "seqPag": 1
            }],  # You'll need to provide this information
        "nfce": {
            "nfceMd5": hashlib.md5(xml_unescaped.encode()).hexdigest(),
            "nfceProt": nfe_data['nProt'],
            "nfceXml": xml_unescaped,
        }
    }
    
    for item in nfe_data['det']:
        det_item = {
            "lCanc": False,
            "lNaoMovEstoque": False,
            "prod": {
                "CFOP": item['prod']['CFOP'],
                "NCM": item['prod']['NCM'],
                "cEAN": item['prod']['cEAN'],
                "cProd": item['prod']['cProd'],
                "cUn": item['prod']['uCom'],
                "nQuant": float(item['prod']['qCom']),
                "vAcresc": 0,
                "vDesc": 0,
                "vItem": float(item['prod']['vProd']),
                "vProd": float(item['prod']['vProd']),
                "vUnit": float(item['prod']['vUnCom']),
                "xProd": item['prod']['xProd']
            },
            "prodIdent": {
                "emiProduto": item['prod']['cProd'],
                "idLocalEstoque": "",  # You'll need to provide this information
                "idProduto": 13  # You'll need to provide this information
            },
            "seqItem": int(item['nItem'])
        }
        omie_json["NFe"]["det"].append(det_item)    
    return omie_json

def build_omie_json(store_name, invoice):
    omie_json = convert_xml_to_omie_json(invoice["xml"])
    
    # Adiciona informações específicas da loja
    omie_json["caixa"]["seqCaixa"] = get_next_sequencial('seqCaixa')
    omie_json["caixa"]["seqCupom"] = get_next_sequencial('seqCupom')
    
    # Define ID do cliente específico para cada loja
    if store_name == 'otro':
        omie_json["cupomIdent"]["idCliente"] = '675944858'
        omie_json["formasPag"][0]["pagIdent"]["idConta"] = 3569457062
    elif store_name == 'tratto':
        omie_json["cupomIdent"]["idCliente"] = '675944859'  # Ajuste este valor conforme necessário
        omie_json["formasPag"][0]["pagIdent"]["idConta"] = 7502625278

    omie_json["emissor"]["emiSerial"] = invoice.get("emiSerial", 1)
    
    return omie_json
    #omie_json["nfce"]["nfceProt"] = nfe_data['nProt']
    
    return omie_json
def create_json_from_omie_json(omie_json, filename=None):
    if filename is None:
        filename = f"omie_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    
    # Cria uma cópia do omie_json para não modificar o original
    json_data = omie_json.copy()
    
    # Adiciona um timestamp ao JSON
    json_data['timestamp'] = datetime.now().isoformat()
    
    # Escreve o JSON em um arquivo
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(json_data, f, ensure_ascii=False, indent=4)
    
    return filename
def process_omie_invoice(store_config, omie_json):
    md5_value = omie_json["nfce"]["nfceMd5"]
    file_name = "processed_nfce_md5.txt"

    # Carregar valores existentes do arquivo
    try:
        with open(file_name, "r") as f:
            processed_md5 = set(f.read().splitlines())
    except FileNotFoundError:
        processed_md5 = set()

    # Verificar se o valor já foi processado
    if md5_value in processed_md5:
        logging.info(f"[{store_config.name}] NF-e já processada (MD5: {md5_value}). Pulando...")
        print(f"[{store_config.name}] NF-e já processada (MD5: {md5_value}). Pulando...")
        return
    time.sleep(180)
    # Fazer a requisição
    url = "https://app.omie.com.br/api/v1/produtos/cupomfiscalincluir/"
    headers = {"Content-Type": "application/json"}
    body = {
        "call": "IncluirNfce",
        "app_key": store_config.omie_app_key,
        "app_secret": store_config.omie_app_secret,
        "param": [omie_json]
    }

    try:
        response = requests.post(url, headers=headers, json=body)
        response_data = response.json()

        if "faultcode" in response_data:
            if response_data["faultcode"] == "SOAP-ENV:Client-3333":
                logging.info(f"[{store_config.name}] Cupom duplicado: {response_data['faultstring']}. Continuando...")
                print(f"[{store_config.name}] Cupom duplicado: {response_data['faultstring']}. Continuando...")
                return
            raise Exception(f"Erro ao processar nota: {response_data['faultstring']}")

        if response.status_code != 200:
            logging.error(f"Unexpected status: {response.status_code}")
            raise Exception(f"Erro ao enviar nota: {response.text}")

        # Salvar MD5 no arquivo após sucesso
        with open(file_name, "a") as f:
            f.write(md5_value + "\n")

        logging.info(f"[{store_config.name}] Nota fiscal enviada com sucesso: {response.text}")
        print(f"[{store_config.name}] Nota fiscal enviada com sucesso: {response.text}")

    except Exception as e:
        logging.error(f"[{store_config.name}] Erro ao enviar nota: {str(e)}")
        print(f"[{store_config.name}] Erro ao enviar nota: {str(e)}")

def execute_zig_omie_integration():
    logging.info("Iniciando integração...")
    
    now = datetime.now()
    last_run = now - timedelta(days=1)

    try:
        with timeout(900):  # Timeout de 4 minutos
            invoices = fetch_invoices(last_run, now, 1)
            for invoice in invoices:
                omie_json = build_omie_json(invoice)
                process_omie_invoice(omie_json)
    except TimeoutError:
        logging.error("O tempo limite foi atingido.")
    except Exception as e:
        logging.error(f"Erro na integração: {e}")

    logging.info("Processamento concluído.")

def parse_nfe_xml(xml_data):
    # Define o namespace
    ns = {'nfe': 'http://www.portalfiscal.inf.br/nfe', 'ds': 'http://www.w3.org/2000/09/xmldsig#'}

    # Função auxiliar para obter o texto de um elemento ou retornar None
    def get_text_or_none(element, tag, ns):
        if element is not None:
            found = element.find(tag, ns)
            return found.text if found is not None else None
        return None

    # Carrega o XML
    root = ET.fromstring(xml_data)

    # Inicializa um dicionário para armazenar os dados
    nfe_data = {}

    # Encontra o elemento 'infNFe' com o namespace
    inf_nfe = root.find('.//nfe:infNFe', ns)
    
    if inf_nfe is not None:
        # Atributos de 'infNFe'
        nfe_data['Id'] = inf_nfe.attrib.get('Id')
        nfe_data['versao'] = inf_nfe.attrib.get('versao')

        # 'ide' - Identificação da NF-e
        ide = inf_nfe.find('nfe:ide', ns)
        if ide is not None:
            nfe_data['ide'] = {
                'cUF': get_text_or_none(ide, 'nfe:cUF', ns),
                'cNF': get_text_or_none(ide, 'nfe:cNF', ns),
                'natOp': get_text_or_none(ide, 'nfe:natOp', ns),
                'mod': get_text_or_none(ide, 'nfe:mod', ns),
                'serie': get_text_or_none(ide, 'nfe:serie', ns),
                'nNF': get_text_or_none(ide, 'nfe:nNF', ns),
                'dhEmi': get_text_or_none(ide, 'nfe:dhEmi', ns),
                'tpNF': get_text_or_none(ide, 'nfe:tpNF', ns),
                'idDest': get_text_or_none(ide, 'nfe:idDest', ns),
                'cMunFG': get_text_or_none(ide, 'nfe:cMunFG', ns),
                'tpImp': get_text_or_none(ide, 'nfe:tpImp', ns),
                'tpEmis': get_text_or_none(ide, 'nfe:tpEmis', ns),
                'cDV': get_text_or_none(ide, 'nfe:cDV', ns),
                'tpAmb': get_text_or_none(ide, 'nfe:tpAmb', ns),
                'finNFe': get_text_or_none(ide, 'nfe:finNFe', ns),
                'indFinal': get_text_or_none(ide, 'nfe:indFinal', ns),
                'indPres': get_text_or_none(ide, 'nfe:indPres', ns),
                'indIntermed': get_text_or_none(ide, 'nfe:indIntermed', ns),
                'procEmi': get_text_or_none(ide, 'nfe:procEmi', ns),
                'verProc': get_text_or_none(ide, 'nfe:verProc', ns)
            }

        # 'emit' - Emitente
        emit = inf_nfe.find('nfe:emit', ns)
        if emit is not None:
            enderEmit = emit.find('nfe:enderEmit', ns)
            nfe_data['emit'] = {
                'CNPJ': get_text_or_none(emit, 'nfe:CNPJ', ns),
                'xNome': get_text_or_none(emit, 'nfe:xNome', ns),
                'xFant': get_text_or_none(emit, 'nfe:xFant', ns),
                'IE': get_text_or_none(emit, 'nfe:IE', ns),
                'CRT': get_text_or_none(emit, 'nfe:CRT', ns),
                'enderEmit': {
                    'xLgr': get_text_or_none(enderEmit, 'nfe:xLgr', ns),
                    'nro': get_text_or_none(enderEmit, 'nfe:nro', ns),
                    'xBairro': get_text_or_none(enderEmit, 'nfe:xBairro', ns),
                    'cMun': get_text_or_none(enderEmit, 'nfe:cMun', ns),
                    'xMun': get_text_or_none(enderEmit, 'nfe:xMun', ns),
                    'UF': get_text_or_none(enderEmit, 'nfe:UF', ns),
                    'CEP': get_text_or_none(enderEmit, 'nfe:CEP', ns),
                    'cPais': get_text_or_none(enderEmit, 'nfe:cPais', ns),
                    'xPais': get_text_or_none(enderEmit, 'nfe:xPais', ns),
                    'fone': get_text_or_none(enderEmit, 'nfe:fone', ns)
                }
            }

        # 'dest' - Destinatário
        dest = inf_nfe.find('nfe:dest', ns)
        if dest is not None:
            nfe_data['dest'] = {
                'CPF': get_text_or_none(dest, 'nfe:CPF', ns),
                'xNome': get_text_or_none(dest, 'nfe:xNome', ns),
                'indIEDest': get_text_or_none(dest, 'nfe:indIEDest', ns)
            }

        # 'det' - Detalhes dos produtos/serviços
        det_list = []
        for det in inf_nfe.findall('nfe:det', ns):
            prod = det.find('nfe:prod', ns)
            imposto = det.find('nfe:imposto', ns)
            det_item = {
                'nItem': det.attrib.get('nItem'),
                'prod': {
                    'cProd': get_text_or_none(prod, 'nfe:cProd', ns),
                    'cEAN': get_text_or_none(prod, 'nfe:cEAN', ns),
                    'xProd': get_text_or_none(prod, 'nfe:xProd', ns),
                    'NCM': get_text_or_none(prod, 'nfe:NCM', ns),
                    'CFOP': get_text_or_none(prod, 'nfe:CFOP', ns),
                    'uCom': get_text_or_none(prod, 'nfe:uCom', ns),
                    'qCom': get_text_or_none(prod, 'nfe:qCom', ns),
                    'vUnCom': get_text_or_none(prod, 'nfe:vUnCom', ns),
                    'vProd': get_text_or_none(prod, 'nfe:vProd', ns),
                    'cEANTrib': get_text_or_none(prod, 'nfe:cEANTrib', ns),
                    'uTrib': get_text_or_none(prod, 'nfe:uTrib', ns),
                    'qTrib': get_text_or_none(prod, 'nfe:qTrib', ns),
                    'vUnTrib': get_text_or_none(prod, 'nfe:vUnTrib', ns),
                    'indTot': get_text_or_none(prod, 'nfe:indTot', ns)
                },
                'imposto': {
                    'vTotTrib': get_text_or_none(imposto, 'nfe:vTotTrib', ns) if imposto is not None else None
                }
            }
            det_list.append(det_item)
        nfe_data['det'] = det_list

        # 'total' - Totais da NF-e
        total = inf_nfe.find('nfe:total', ns)
        if total is not None:
            icms_tot = total.find('nfe:ICMSTot', ns)
            if icms_tot is not None:
                nfe_data['total'] = {
                    'vBC': get_text_or_none(icms_tot, 'nfe:vBC', ns),
                    'vICMS': get_text_or_none(icms_tot, 'nfe:vICMS', ns),
                    'vICMSDeson': get_text_or_none(icms_tot, 'nfe:vICMSDeson', ns),
                    'vFCP': get_text_or_none(icms_tot, 'nfe:vFCP', ns),
                    'vBCST': get_text_or_none(icms_tot, 'nfe:vBCST', ns),
                    'vST': get_text_or_none(icms_tot, 'nfe:vST', ns),
                    'vFCPST': get_text_or_none(icms_tot, 'nfe:vFCPST', ns),
                    'vFCPSTRet': get_text_or_none(icms_tot, 'nfe:vFCPSTRet', ns),
                    'vProd': get_text_or_none(icms_tot, 'nfe:vProd', ns),
                    'vFrete': get_text_or_none(icms_tot, 'nfe:vFrete', ns),
                    'vSeg': get_text_or_none(icms_tot, 'nfe:vSeg', ns),
                    'vDesc': get_text_or_none(icms_tot, 'nfe:vDesc', ns),
                    'vII': get_text_or_none(icms_tot, 'nfe:vII', ns),
                    'vIPI': get_text_or_none(icms_tot, 'nfe:vIPI', ns),
                    'vIPIDevol': get_text_or_none(icms_tot, 'nfe:vIPIDevol', ns),
                    'vPIS': get_text_or_none(icms_tot, 'nfe:vPIS', ns),
                    'vCOFINS': get_text_or_none(icms_tot, 'nfe:vCOFINS', ns),
                    'vOutro': get_text_or_none(icms_tot, 'nfe:vOutro', ns),
                    'vNF': get_text_or_none(icms_tot, 'nfe:vNF', ns),
                    'vTotTrib': get_text_or_none(icms_tot, 'nfe:vTotTrib', ns)
                }

        # 'infProt' - Protocolo de Autorização
        prot_nfe = root.find('.//nfe:infProt', ns)
        if prot_nfe is not None:
            nfe_data['nProt'] = get_text_or_none(prot_nfe, 'nfe:nProt', ns)

    else:
        print("Elemento 'infNFe' não encontrado no XML.")

    # Converte o dicionário para JSON
    nfe_json = json.dumps(nfe_data, indent=4, ensure_ascii=False)
    return nfe_json, nfe_data  # Retorna o JSON e o dicionário de dados
    
def get_next_sequencial(tipo):
    filename = 'sequenciais.json'
    today = datetime.now().strftime('%Y-%m-%d')
    
    try:
        with open(filename, 'r') as f:
            data = json.load(f)
    except FileNotFoundError:
        data = {}

    if today not in data:
        data[today] = {'seqCaixa': 0, 'seqCupom': 0}

    data[today][tipo] += 1
    next_seq = data[today][tipo]

    with open(filename, 'w') as f:
        json.dump(data, f)

    return next_seq
def execute_store_integration(store_name):
    store_config = config.stores[store_name]
    logging.info(f"Iniciando integração para loja {store_name}...")
    
    now = datetime.now()
    last_run = now - timedelta(days=1)
    try:
        with timeout(900):  # Timeout de 4 minutos
            invoices = fetch_invoices(store_config, last_run, now, 1)
            for invoice in invoices:
                omie_json = build_omie_json(store_name, invoice)
                process_omie_invoice(store_config, omie_json)

            logging.info(f"[{store_name}] Remessa de vendas finalizada. Aguardando a próxima remessa em algumas horas.")
            print(f"[{store_name}] Remessa de vendas finalizada. Aguardando a próxima remessa em algumas horas.")
    except TimeoutError:
        logging.error(f"[{store_name}] O tempo limite foi atingido.")
    except Exception as e:
        logging.error(f"[{store_name}] Erro na integração: {e}")
        print(f"[{store_name}] Erro na integração: {e}")
    finally:
        logging.info(f"[{store_name}] Processamento concluído.")

def execute_all_integrations():
    for store_name in config.stores:
        execute_store_integration(store_name)

if __name__ == "__main__":
    # Configurar logging para cada loja
    for store_name in config.stores:
        logging.basicConfig(
            filename=f'integration_{store_name}.log',
            level=logging.INFO,
            format=f'%(asctime)s [%(levelname)s] [{store_name}] %(message)s'
        )
    
    # Executa a integração imediatamente
    logging.info("Executando integração imediatamente...")
    execute_all_integrations()

    # Agenda a execução periódica
    scheduler.add_job(execute_all_integrations, 'interval', seconds=21600)
    scheduler.start()
